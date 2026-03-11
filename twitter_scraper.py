"""
Twitter/X Web Scraper com Playwright
=====================================
Coleta publicações de um usuário do X (Twitter) e salva em XLSX.

Requisitos:
    pip install playwright openpyxl
    playwright install chromium

Uso:
    python twitter_scraper.py --usuario elonmusk --quantidade 20
    python twitter_scraper.py --usuario elonmusk --quantidade 50 --saida meus_tweets.xlsx
"""

import asyncio
import re
import argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def limpar_numero(texto: str) -> int:
    """Converte strings como '1,2K' ou '3M' em inteiros."""
    if not texto or texto.strip() in ("", "—", "-"):
        return 0
    texto = texto.strip().replace(",", ".")
    try:
        if texto.endswith("K") or texto.endswith("k"):
            return int(float(texto[:-1]) * 1_000)
        if texto.endswith("M") or texto.endswith("m"):
            return int(float(texto[:-1]) * 1_000_000)
        return int(texto.replace(".", ""))
    except ValueError:
        return 0


def salvar_xlsx(tweets: list[dict], caminho: str) -> None:
    if not tweets:
        print("⚠️  Nenhum tweet para salvar.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tweets"

    # Cabeçalho
    colunas = ["#", "Autor", "Usuário", "Descrição", "Data", "Visualizações", "Likes", "Comentários", "Retweets", "Link"]
    header_fill = PatternFill("solid", fgColor="1DA1F2")
    header_font = Font(bold=True, color="FFFFFF")

    for col, titulo in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Larguras das colunas
    larguras = [5, 20, 18, 60, 18, 15, 10, 14, 12, 50]
    for col, largura in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = largura

    # Linhas de dados
    fill_par   = PatternFill("solid", fgColor="E8F5FE")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    for i, t in enumerate(tweets, 1):
        fill = fill_par if i % 2 == 0 else fill_impar
        linha = [
            i,
            t["autor"],
            t["usuario"],
            t["descricao"],
            t["data"],
            t["visualizacoes"],
            t["likes"],
            t["comentarios"],
            t["retweets"],
            t["url_tweet"],
        ]
        for col, valor in enumerate(linha, 1):
            cell = ws.cell(row=i + 1, column=col, value=valor)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=(col == 4), vertical="top")

    ws.freeze_panes = "A2"
    wb.save(caminho)
    print(f"✅  {len(tweets)} tweets salvos em: {caminho}")


# ──────────────────────────────────────────────
# Scraper principal
# ──────────────────────────────────────────────

async def coletar_tweets(usuario: str, quantidade: int = 20) -> list[dict]:
    """
    Abre o perfil público do usuário no X e coleta os tweets visíveis.
    Faz scroll até atingir `quantidade` tweets ou não haver mais conteúdo.
    """
    url_perfil = f"https://x.com/{usuario}"
    tweets_coletados: list[dict] = []
    ids_vistos: set[str] = set()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0.0.0 Safari/537.36"
            ),
            locale="pt-BR",
        )
        page = await context.new_page()

        print(f"🌐  Acessando perfil: {url_perfil}")
        try:
            await page.goto(url_perfil, wait_until="domcontentloaded", timeout=30_000)
        except PlaywrightTimeout:
            print("⚠️  Tempo limite ao carregar a página.")
            await browser.close()
            return []

        # Aguarda os tweets aparecerem
        try:
            await page.wait_for_selector("article[data-testid='tweet']", timeout=20_000)
        except PlaywrightTimeout:
            print("⚠️  Nenhum tweet encontrado. O perfil pode ser privado ou o X bloqueou o acesso.")
            await browser.close()
            return []

        tentativas_sem_novos = 0
        MAX_TENTATIVAS = 5

        while len(tweets_coletados) < quantidade and tentativas_sem_novos < MAX_TENTATIVAS:
            artigos = await page.query_selector_all("article[data-testid='tweet']")
            novos = 0

            for artigo in artigos:
                # ── ID único via link do tweet ──────────────────────────
                link_el = await artigo.query_selector("a[href*='/status/']")
                if not link_el:
                    continue
                href = await link_el.get_attribute("href") or ""
                match = re.search(r"/status/(\d+)", href)
                tweet_id = match.group(1) if match else href
                if tweet_id in ids_vistos:
                    continue
                ids_vistos.add(tweet_id)

                # ── Autor / @ ───────────────────────────────────────────
                autor = ""
                usuario_handle = usuario
                nome_el = await artigo.query_selector("[data-testid='User-Name']")
                if nome_el:
                    spans = await nome_el.query_selector_all("span")
                    textos = [await s.inner_text() for s in spans]
                    textos = [t.strip() for t in textos if t.strip()]
                    for t in textos:
                        if t.startswith("@"):
                            usuario_handle = t[1:]
                        elif not autor:
                            # Remove emojis e caracteres corrompidos, mantém letras latinas
                            limpo = re.sub(r'[^\w\s\-\.À-ɏ]', '', t).strip()
                            if limpo:
                                autor = limpo
                    if not autor:
                        autor = usuario_handle

                # ── Texto do tweet ──────────────────────────────────────
                descricao = ""
                texto_el = await artigo.query_selector("[data-testid='tweetText']")
                if texto_el:
                    descricao = (await texto_el.inner_text()).strip()

                # ── Data ────────────────────────────────────────────────
                data = ""
                time_el = await artigo.query_selector("time")
                if time_el:
                    data = await time_el.get_attribute("datetime") or ""
                    # Formata para DD/MM/YYYY HH:MM
                    try:
                        dt = datetime.fromisoformat(data.replace("Z", "+00:00"))
                        data = dt.strftime("%d/%m/%Y %H:%M")
                    except ValueError:
                        pass

                # ── Métricas (likes, comentários, retweets, views) ──────
                async def get_metric(testid: str) -> int:
                    el = await artigo.query_selector(f"[data-testid='{testid}']")
                    if el:
                        txt = await el.get_attribute("aria-label") or await el.inner_text()
                        nums = re.findall(r"[\d.,]+[KkMm]?", txt)
                        return limpar_numero(nums[0]) if nums else 0
                    return 0

                comentarios  = await get_metric("reply")
                retweets     = await get_metric("retweet")
                likes        = await get_metric("like")

                # Views ficam num span próximo ao ícone de gráfico
                visualizacoes = 0
                views_els = await artigo.query_selector_all("a[href*='/analytics'] span")
                for el in views_els:
                    txt = (await el.inner_text()).strip()
                    if txt:
                        visualizacoes = limpar_numero(txt)
                        break

                url_tweet = f"https://x.com/{usuario_handle}/status/{tweet_id}"

                tweets_coletados.append({
                    "autor":         autor,
                    "usuario":       f"@{usuario_handle}",
                    "descricao":     descricao,
                    "data":          data,
                    "visualizacoes": visualizacoes,
                    "likes":         likes,
                    "comentarios":   comentarios,
                    "retweets":      retweets,
                    "url_tweet":     url_tweet,
                })
                novos += 1

                if len(tweets_coletados) >= quantidade:
                    break

            print(f"   → {len(tweets_coletados)}/{quantidade} tweets coletados…")

            if novos == 0:
                tentativas_sem_novos += 1
            else:
                tentativas_sem_novos = 0

            if len(tweets_coletados) < quantidade:
                await page.evaluate("window.scrollBy(0, 2000)")
                await page.wait_for_timeout(2000)

        await browser.close()

    return tweets_coletados[:quantidade]


# ──────────────────────────────────────────────
# Entrypoint
# ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Coleta tweets de um usuário do X (Twitter) usando Playwright."
    )
    parser.add_argument(
        "--usuario", "-u",
        required=True,
        help="Nome de usuário do X sem o '@'. Ex: elonmusk"
    )
    parser.add_argument(
        "--quantidade", "-q",
        type=int,
        default=20,
        help="Número máximo de tweets a coletar (padrão: 20)"
    )
    parser.add_argument(
        "--saida", "-o",
        default="",
        help="Caminho do arquivo XLSX de saída (padrão: tweets_<usuario>.xlsx)"
    )
    args = parser.parse_args()

    arquivo_saida = args.saida or f"tweets_{args.usuario}.xlsx"

    tweets = asyncio.run(coletar_tweets(args.usuario, args.quantidade))

    if tweets:
        salvar_xlsx(tweets, arquivo_saida)
        print("\n📊  Prévia dos primeiros registros:")
        print(f"{'AUTOR':<20} {'DATA':<18} {'LIKES':>6} {'VIEWS':>8}  TEXTO")
        print("─" * 80)
        for t in tweets[:5]:
            texto_curto = (t["descricao"] or "")[:35].replace("\n", " ")
            print(
                f"{t['autor']:<20} {t['data']:<18} "
                f"{t['likes']:>6} {t['visualizacoes']:>8}  {texto_curto}…"
            )
    else:
        print("❌  Nenhum tweet foi coletado.")


if __name__ == "__main__":
    main()